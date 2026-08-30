import sys
sys.stdout.reconfigure(encoding="utf-8")
# Valkyrie Quant Desk - 360 Degree Quant Brain Module
import openpyxl, json, os, sys
from datetime import datetime

class ValkyrieQuantBrain:
    def __init__(self, excel_path=None, json_path=None):
        self.excel_path = excel_path
        self.json_path = json_path
        self.history = []
        self.balance = 100000.0
        self.load_data()

    def load_data(self):
        if self.json_path and os.path.exists(self.json_path):
            with open(self.json_path, 'r', encoding='utf-8') as f:
                d = json.load(f)
                self.history = d.get('history', [])
                self.balance = d.get('balance', 100000.0)
        elif self.excel_path and os.path.exists(self.excel_path):
            wb = openpyxl.load_workbook(self.excel_path, data_only=True)
            if '📜 DETAYLI İŞLEM DEFTERİ' in wb.sheetnames:
                ws = wb['📜 DETAYLI İŞLEM DEFTERİ']
                headers = [c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
                for r in ws.iter_rows(min_row=2, values_only=True):
                    if r[0]:
                        self.history.append(dict(zip(headers, r)))

    def run_360_audit(self):
        if not self.history:
            return {'error': 'Analiz edilecek veri bulunamadı!'}

        total_trades = len(self.history)
        wins = [h for h in self.history if float(h.get('net_pnl', h.get('Net PnL ($)', 0.0)) or 0.0) >= 0]
        losses = [h for h in self.history if float(h.get('net_pnl', h.get('Net PnL ($)', 0.0)) or 0.0) < 0]

        total_gross = sum(float(h.get('gross_pnl', h.get('Brüt PnL ($)', 0.0)) or 0.0) for h in self.history)
        total_fees = sum(float(h.get('fees', h.get('Komisyon ($)', 0.0)) or 0.0) for h in self.history)
        total_net = sum(float(h.get('net_pnl', h.get('Net PnL ($)', 0.0)) or 0.0) for h in self.history)

        win_pnl = sum(float(h.get('net_pnl', h.get('Net PnL ($)', 0.0)) or 0.0) for h in wins)
        loss_pnl = sum(float(h.get('net_pnl', h.get('Net PnL ($)', 0.0)) or 0.0) for h in losses)

        avg_win = win_pnl / len(wins) if wins else 0.0
        avg_loss = abs(loss_pnl) / len(losses) if losses else 0.0
        payoff_ratio = (avg_win / avg_loss) if avg_loss > 0 else 0.0
        profit_factor = (win_pnl / abs(loss_pnl)) if loss_pnl != 0 else 0.0
        win_rate = (len(wins) / total_trades * 100.0) if total_trades > 0 else 0.0

        # 1. Temporal / Süre
        dur_buckets = {'0-15dk (Şimşek)': {'w': 0, 'l': 0}, '15-60dk (Hızlı)': {'w': 0, 'l': 0}, '1-3saat (Standart)': {'w': 0, 'l': 0}, '>3saat (Trend)': {'w': 0, 'l': 0}}
        for h in self.history:
            pnl = float(h.get('net_pnl', h.get('Net PnL ($)', 0.0)) or 0.0)
            dur_str = str(h.get('duration', h.get('Süre', '30dk')))
            mins = 30
            if 'sa' in dur_str:
                parts = dur_str.split('sa')
                mins = int(parts[0].strip()) * 60 + (int(parts[1].replace('dk', '').strip()) if len(parts)>1 and parts[1].strip() else 0)
            elif 'dk' in dur_str:
                mins = int(dur_str.replace('dk', '').strip() or 30)

            b = '0-15dk (Şimşek)' if mins <= 15 else ('15-60dk (Hızlı)' if mins <= 60 else ('1-3saat (Standart)' if mins <= 180 else '>3saat (Trend)'))
            if pnl >= 0: dur_buckets[b]['w'] += 1
            else: dur_buckets[b]['l'] += 1

        # 2. Coin Persona
        pair_stats = {}
        for h in self.history:
            sym = str(h.get('symbol', h.get('Parite', 'Bilinmeyen')))
            pnl = float(h.get('net_pnl', h.get('Net PnL ($)', 0.0)) or 0.0)
            fee = float(h.get('fees', h.get('Komisyon ($)', 0.0)) or 0.0)
            if sym not in pair_stats:
                pair_stats[sym] = {'pnl': 0.0, 'fees': 0.0, 'wins': 0, 'losses': 0, 'trades': 0}
            pair_stats[sym]['pnl'] += pnl
            pair_stats[sym]['fees'] += fee
            pair_stats[sym]['trades'] += 1
            if pnl >= 0: pair_stats[sym]['wins'] += 1
            else: pair_stats[sym]['losses'] += 1

        sorted_coins = sorted(pair_stats.items(), key=lambda x: x[1]['pnl'], reverse=True)
        champions = sorted_coins[:7]
        toxic = sorted_coins[-7:]

        # 3. Setup Lab
        setup_stats = {}
        for h in self.history:
            st = str(h.get('setup_id', h.get('reason', h.get('Giriş Stratejisi / Formasyon', 'Genel')))).split('(')[0].strip()
            pnl = float(h.get('net_pnl', h.get('Net PnL ($)', 0.0)) or 0.0)
            if st not in setup_stats:
                setup_stats[st] = {'pnl': 0.0, 'wins': 0, 'losses': 0, 'trades': 0}
            setup_stats[st]['pnl'] += pnl
            setup_stats[st]['trades'] += 1
            if pnl >= 0: setup_stats[st]['wins'] += 1
            else: setup_stats[st]['losses'] += 1

        return {
            'total_trades': total_trades,
            'win_rate': win_rate,
            'wins_count': len(wins),
            'loss_count': len(losses),
            'total_gross': total_gross,
            'total_fees': total_fees,
            'total_net': total_net,
            'avg_win': avg_win,
            'avg_loss': avg_loss,
            'payoff_ratio': payoff_ratio,
            'profit_factor': profit_factor,
            'duration_buckets': dur_buckets,
            'champions': champions,
            'toxic': toxic,
            'setup_stats': setup_stats
        }

    def print_360_report(self):
        r = self.run_360_audit()
        if 'error' in r:
            print(r['error'])
            return

        print('=' * 85)
        print('🧠 VALKYRIE QUANT BRAIN — 360° OTOMATİK ADLİ ANALİZ RAPORU')
        print(f'Oluşturulma: {datetime.now().strftime("%d.%m.%Y %H:%M:%S")} | Toplam İncelenen İşlem: {r["total_trades"]}')
        print('=' * 85)

        print('\n📊 1. PERFORMANS & KÂRLILIK MATRİSİ:')
        print(f'   • Win Rate (Kazanma Oranı) : %{r["win_rate"]:.1f} ({r["wins_count"]} Win / {r["loss_count"]} Loss)')
        print(f'   • Toplam Brüt PnL          : ${r["total_gross"]:+,.2f} USDT')
        print(f'   • Toplam Borsa Komisyonu   : ${r["total_fees"]:,.2f} USDT')
        print(f'   • Kasaya Kalan Net PnL     : ${r["total_net"]:+,.2f} USDT')
        print(f'   • Ortalama Kazanç / Kayıp  : +${r["avg_win"]:.2f} / -${r["avg_loss"]:.2f}')
        print(f'   • ⚖️ Payoff Ratio           : {r["payoff_ratio"]:.2f}x (Hedef: > 1.30x)')
        print(f'   • 🏆 Profit Factor         : {r["profit_factor"]:.2f}')

        print('\n⏱️ 2. ZAMAN & SÜRE (HOLD DURATION) BOYUTU:')
        for k, v in r['duration_buckets'].items():
            tot = v['w'] + v['l']
            wr = v['w'] / tot * 100 if tot > 0 else 0
            print(f'   • {k:20s}: {v["w"]:3d}W / {v["l"]:3d}L (%{wr:4.1f} WR) | Toplam: {tot:3d} İşlem')

        print('\n👑 3. EN ÇOK KAZANDIRAN COINLER (KASA MOTORLARI):')
        for sym, d in r['champions']:
            wr = d['wins'] / d['trades'] * 100 if d['trades'] > 0 else 0
            print(f'   • {sym:12s}: +${d["pnl"]:+,.2f} USDT ({d["wins"]:2d}W / {d["losses"]:2d}L - %{wr:.0f} WR) | Komisyon: ${d["fees"]:.2f}')

        print('\n🚨 4. EN ÇOK KAYBETTİREN COINLER (KASA DELİKLERİ):')
        for sym, d in r['toxic']:
            wr = d['wins'] / d['trades'] * 100 if d['trades'] > 0 else 0
            print(f'   • {sym:12s}: -${abs(d["pnl"]):6.2f} USDT ({d["wins"]:2d}W / {d["losses"]:2d}L - %{wr:.0f} WR) | Komisyon: ${d["fees"]:.2f}')

        print(f'\n🔬 5. SETUP BAZINDA QUANT LAB PERFORMANSI:')
        for st, d in sorted(r['setup_stats'].items(), key=lambda x: x[1]['pnl'], reverse=True):
            wr = d['wins'] / d['trades'] * 100 if d['trades'] > 0 else 0
            print(f'   • {st:35s}: ${d["pnl"]:+,.2f} USDT | {d["wins"]:2d}W / {d["losses"]:2d}L (%{wr:4.1f} WR)')

        # 6. Görsel ve Adli Grafik Delilleri
        charts_dir = 'c:/Users/aucar/Desktop/trade-bot/ANALİZ/KRİTİK_GRAFİKLER'
        archived_charts = []
        if os.path.exists(charts_dir):
            archived_charts = [f for f in os.listdir(charts_dir) if f.endswith('.png') or f.endswith('.jpeg') or f.endswith('.jpg')]

        print(f'\n🖼️ 6. ADLİ GÖRSEL DELİL ARŞİVİ ({len(archived_charts)} Adet Kritik Grafik Kayıtlı):')
        if archived_charts:
            for f_name in sorted(archived_charts, reverse=True)[:6]:
                print(f'   • 📸 {f_name}')
        else:
            print('   • Henüz yeni arşivlenmiş kritik grafik yok (Otomatik yakalayıcı devrede).')

        print('=' * 85)

if __name__ == '__main__':
    brain = ValkyrieQuantBrain(json_path='c:/Users/aucar/Desktop/trade-bot/trade_history.json')
    brain.print_360_report()
